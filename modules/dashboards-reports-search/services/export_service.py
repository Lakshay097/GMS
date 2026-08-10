"""
Export pipeline — R-59/BR-17.

Formats: Excel (.xlsx via openpyxl), CSV (stdlib), PDF (fpdf2), API (JSON).

Category-level export/view restrictions (BR-04/BR-19/R-50):
  Before any export the service queries ``kpi_category_export_restrictions``
  for the caller's role.  If the report would expose rows whose category_code
  is restricted for this role, the export is DENIED with AuthorizationError.
  SuperAdmin is exempt from category restrictions.

Architecture note:
  The export pipeline is intentionally on the READ path:
  * Queries go through the read-replica session (never the write engine).
  * File generation happens in the same request for CSV/API (fast, small).
  * Excel/PDF are also synchronous here but the API enqueues a
    report_export_jobs row first so the caller can poll for completion.
    In production a worker process would pick up 'pending' jobs; for Phase 1
    the worker runs in-process (acceptable for single-node deployment).
"""
from __future__ import annotations

import csv
import io
import json
import logging
import os
import uuid
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional
from uuid import UUID

from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from modules.dashboards_reports_search.schemas import (
    ExportJobResponse,
    ExportRequest,
    ReportFilter,
)
from modules.dashboards_reports_search.services.report_service import ReportService
from shared.errors import AuthorizationError, ValidationError
from shared.middleware.tenancy import TenantContext

logger = logging.getLogger(__name__)

EXPORT_URL_PREFIX = os.getenv("EXPORT_URL_PREFIX", "/api/v1/exports/files/")


# ── Category restriction check ────────────────────────────────────────────────

async def _check_category_restrictions(
    db: AsyncSession,
    roles: List[str],
    rows: List[Dict[str, Any]],
) -> None:
    """
    Raise AuthorizationError if any row in the export contains a category_code
    that is restricted for the caller's role (BR-04/BR-19/R-50).
    SuperAdmin is unconditionally exempt.
    """
    lower_roles = [r.lower() for r in roles]
    if "superadmin" in lower_roles:
        return

    # Collect unique category codes present in the result set
    category_codes = {
        row.get("category_code")
        for row in rows
        if row.get("category_code")
    }
    if not category_codes:
        return

    for role in lower_roles:
        # Use expandable IN instead of ANY() to stay compatible with SQLite (tests)
        # and PostgreSQL (production).
        if not category_codes:
            continue
        placeholders = ", ".join(f":code_{i}" for i, _ in enumerate(category_codes))
        code_params = {f"code_{i}": c for i, c in enumerate(category_codes)}
        result = await db.execute(
            text(
                f"""
                SELECT category_code
                FROM kpi_category_export_restrictions
                WHERE category_code IN ({placeholders})
                  AND restricted_role = :role
                  AND restrict_export = 1
                """
            ),
            {**code_params, "role": role},
        )
        blocked = [r.category_code for r in result.fetchall()]
        if blocked:
            raise AuthorizationError(
                f"Export denied: category(s) {blocked} are restricted for role '{role}' "
                f"(BR-04/BR-19/R-50). Contact your administrator."
            )


async def _check_view_restrictions(
    db: AsyncSession,
    roles: List[str],
    rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """
    Strip rows whose category_code is view-restricted for the caller's role.
    Returns the filtered row list.
    """
    lower_roles = [r.lower() for r in roles]
    if "superadmin" in lower_roles:
        return rows

    category_codes = {
        row.get("category_code")
        for row in rows
        if row.get("category_code")
    }
    if not category_codes:
        return rows

    restricted: set[str] = set()
    for role in lower_roles:
        if not category_codes:
            break
        placeholders = ", ".join(f":code_{i}" for i, _ in enumerate(category_codes))
        code_params = {f"code_{i}": c for i, c in enumerate(category_codes)}
        result = await db.execute(
            text(
                f"""
                SELECT category_code
                FROM kpi_category_export_restrictions
                WHERE category_code IN ({placeholders})
                  AND restricted_role = :role
                  AND restrict_view = 1
                """
            ),
            {**code_params, "role": role},
        )
        restricted.update(r.category_code for r in result.fetchall())

    if not restricted:
        return rows
    return [r for r in rows if r.get("category_code") not in restricted]


# ── Format renderers ──────────────────────────────────────────────────────────

def _render_csv(rows: List[Dict[str, Any]]) -> bytes:
    if not rows:
        return b""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    for row in rows:
        writer.writerow({k: str(v) if v is not None else "" for k, v in row.items()})
    return buf.getvalue().encode("utf-8-sig")


def _render_excel(rows: List[Dict[str, Any]], report_type: str) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise ValidationError(
            "Excel export requires openpyxl. Install it with: pip install openpyxl"
        )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_type[:31]  # Excel sheet name max 31 chars
    if rows:
        headers = list(rows[0].keys())
        header_fill = PatternFill("solid", fgColor="003366")
        header_font = Font(color="FFFFFF", bold=True)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
        for row_idx, row in enumerate(rows, 2):
            for col_idx, key in enumerate(headers, 1):
                val = row.get(key)
                if isinstance(val, (dict, list)):
                    val = json.dumps(val)
                ws.cell(row=row_idx, column=col_idx, value=val)
        # Auto-size columns (capped at 60)
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _render_pdf(rows: List[Dict[str, Any]], report_type: str) -> bytes:
    try:
        from fpdf import FPDF
    except ImportError:
        raise ValidationError(
            "PDF export requires fpdf2. Install it with: pip install fpdf2"
        )
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, report_type.replace("_", " ").title(), ln=True)
    pdf.set_font("Helvetica", "", 8)
    pdf.cell(0, 6, f"Generated: {datetime.now(timezone.utc).isoformat()}", ln=True)
    pdf.ln(4)
    if rows:
        headers = list(rows[0].keys())
        col_w = min(270 / max(len(headers), 1), 60)
        pdf.set_font("Helvetica", "B", 7)
        pdf.set_fill_color(0, 51, 102)
        pdf.set_text_color(255, 255, 255)
        for h in headers:
            pdf.cell(col_w, 7, str(h)[:20], border=1, fill=True)
        pdf.ln()
        pdf.set_font("Helvetica", "", 7)
        pdf.set_text_color(0, 0, 0)
        for i, row in enumerate(rows):
            pdf.set_fill_color(240, 245, 255) if i % 2 == 0 else pdf.set_fill_color(255, 255, 255)
            for key in headers:
                val = row.get(key)
                if isinstance(val, (dict, list)):
                    val = json.dumps(val)
                pdf.cell(col_w, 6, str(val or "")[:25], border=1, fill=True)
            pdf.ln()
    return pdf.output()



# ── Export service ────────────────────────────────────────────────────────────

class ExportService:
    """
    Coordinates report data retrieval → category restriction check → format render.
    Uses the read-replica session for all data queries (R-61).
    """

    def __init__(self, read_db: AsyncSession, write_db: AsyncSession) -> None:
        self.read_db = read_db    # for report queries
        self.write_db = write_db  # for writing export job rows

    async def enqueue_and_run(
        self,
        body: ExportRequest,
        tenant: TenantContext,
    ) -> ExportJobResponse:
        """
        Enqueue a report_export_jobs row, run the export synchronously (Phase 1),
        and update the job status.  Returns the completed job descriptor.
        """
        job_id = uuid.uuid4()
        now = datetime.now(timezone.utc)
        flt = body.filters or ReportFilter()

        # Write the job record (write pool — lightweight single-row insert)
        await self.write_db.execute(
            text(
                """
                INSERT INTO report_export_jobs
                    (id, requested_by, school_id, report_type, format, filters,
                     status, enqueued_at)
                VALUES
                    (:id, :user, :school_id, :report_type, :format, :filters,
                     'pending', :now)
                """
            ),
            {
                "id": str(job_id),
                "user": tenant.user_id,
                "school_id": tenant.school_id,
                "report_type": body.report_type,
                "format": body.format,
                "filters": json.dumps(flt.model_dump(mode="json")),
                "now": now,
            },
        )
        await self.write_db.commit()

        # Mark processing
        await self.write_db.execute(
            text("UPDATE report_export_jobs SET status='processing', started_at=:now WHERE id=:id"),
            {"now": now, "id": str(job_id)},
        )
        await self.write_db.commit()

        try:
            result_url, row_count, file_size = await self._execute_export(
                body, flt, tenant, job_id
            )
            await self.write_db.execute(
                text(
                    """
                    UPDATE report_export_jobs
                    SET status='completed', completed_at=:now,
                        result_url=:url, row_count=:rc, file_size_bytes=:fs
                    WHERE id=:id
                    """
                ),
                {"now": datetime.now(timezone.utc), "url": result_url,
                 "rc": row_count, "fs": file_size, "id": str(job_id)},
            )
            await self.write_db.commit()
            return ExportJobResponse(
                job_id=job_id,
                status="completed",
                report_type=body.report_type,
                format=body.format,
                enqueued_at=now,
                result_url=result_url,
                completed_at=datetime.now(timezone.utc),
                row_count=row_count,
                file_size_bytes=file_size,
            )
        except AuthorizationError:
            await self.write_db.execute(
                text("UPDATE report_export_jobs SET status='failed', error_detail=:err WHERE id=:id"),
                {"err": "Category export restriction denied", "id": str(job_id)},
            )
            await self.write_db.commit()
            raise
        except Exception as exc:
            logger.error("Export job %s failed: %s", job_id, exc)
            await self.write_db.execute(
                text("UPDATE report_export_jobs SET status='failed', error_detail=:err WHERE id=:id"),
                {"err": str(exc)[:1000], "id": str(job_id)},
            )
            await self.write_db.commit()
            raise

    async def get_job(self, job_id: UUID, tenant: TenantContext) -> ExportJobResponse:
        result = await self.write_db.execute(
            text(
                """
                SELECT id, status, report_type, format, enqueued_at,
                       result_url, completed_at, error_detail, row_count, file_size_bytes
                FROM report_export_jobs
                WHERE id = :id AND requested_by = :user
                """
            ),
            {"id": str(job_id), "user": tenant.user_id},
        )
        row = result.fetchone()
        if not row:
            from shared.errors import NotFoundError
            raise NotFoundError("ExportJob")
        return ExportJobResponse(
            job_id=UUID(str(row.id)),
            status=row.status,
            report_type=row.report_type,
            format=row.format,
            enqueued_at=row.enqueued_at,
            result_url=row.result_url,
            completed_at=row.completed_at,
            error_detail=row.error_detail,
            row_count=row.row_count,
            file_size_bytes=row.file_size_bytes,
        )

    # ── Private ────────────────────────────────────────────────────────────────

    async def _execute_export(
        self,
        body: ExportRequest,
        flt: ReportFilter,
        tenant: TenantContext,
        job_id: UUID,
    ) -> tuple[str, int, int]:
        """Run the report query, check restrictions, render, return (url, rows, bytes)."""
        svc = ReportService(self.read_db)
        report = await svc.run(body.report_type, flt, tenant)
        rows = report.rows

        # Apply view-level category restrictions first (strips rows)
        rows = await _check_view_restrictions(self.read_db, tenant.roles, rows)

        # Then check export-level restrictions (raises if blocked)
        await _check_category_restrictions(self.read_db, tenant.roles, rows)

        if body.format == "api":
            payload = json.dumps({"data": rows, "total": len(rows)}).encode()
            file_size = len(payload)
            # Store inline in result_url as a data URI for API format
            result_url = f"{EXPORT_URL_PREFIX}{job_id}.json"
        elif body.format == "csv":
            payload = _render_csv(rows)
            file_size = len(payload)
            result_url = f"{EXPORT_URL_PREFIX}{job_id}.csv"
        elif body.format == "excel":
            payload = _render_excel(rows, body.report_type)
            file_size = len(payload)
            result_url = f"{EXPORT_URL_PREFIX}{job_id}.xlsx"
        elif body.format == "pdf":
            payload = _render_pdf(rows, body.report_type)
            file_size = len(payload)
            result_url = f"{EXPORT_URL_PREFIX}{job_id}.pdf"
        else:
            raise ValidationError(f"Unsupported export format: {body.format}")

        # In production upload to Cloudinary/S3; for Phase 1 store in memory cache
        # keyed by job_id and serve from the /exports/files/{job_id}.{ext} endpoint.
        _export_cache[str(job_id)] = (body.format, payload)

        return result_url, len(rows), file_size


# In-process file cache — replaced by cloud storage in production
_export_cache: Dict[str, tuple[str, bytes]] = {}


def get_export_file(job_id: str) -> Optional[tuple[str, bytes]]:
    """Retrieve a rendered export file by job ID (format, raw_bytes)."""
    return _export_cache.get(job_id)
